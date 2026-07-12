import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils.encoding import force_str

class ExcelExportService:
    @staticmethod
    def generate_excel(data: list, columns: list, title: str = "Rapor") -> bytes:
        """
        Generates an Excel file from a list of dictionaries or lists.
        :param data: List of rows (dicts or lists)
        :param columns: List of column names or dicts with {key, label, width}
        :param title: Sheet title
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title[:30] # Excel limit

        # Header style
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Prepare headers
        headers = []
        keys = []
        for col in columns:
            if isinstance(col, dict):
                headers.append(force_str(col.get('label', col.get('key'))))
                keys.append(col.get('key'))
            else:
                headers.append(force_str(col))
                keys.append(str(col))

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, key in enumerate(keys, 1):
                value = ""
                if isinstance(row_data, dict):
                    value = row_data.get(key, "")
                elif isinstance(row_data, (list, tuple)):
                    # if col_num-1 is within index
                    if col_num - 1 < len(row_data):
                        value = row_data[col_num - 1]
                
                if not isinstance(value, (int, float, bool, type(None))):
                    value = force_str(value)
                
                ws.cell(row=row_num, column=col_num, value=value)

        # Autofit column widths (rough estimate)
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
